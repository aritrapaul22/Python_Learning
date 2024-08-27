from os.path import abspath
import re
import openpyxl
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Color


f_path = abspath("resources/transactions.html")

html = open(f_path).read()
soup = BeautifulSoup(html)
tables = soup.find_all("table")
wb = Workbook()
ws = wb.active
ws.title = 'Example'

iRowCounter = 1
sample_table_headers = ['Transaction Name', 'Min(In secs)', 'Avg(In secs)', '90 Percentile(In secs)', 'Max(In secs)']
for table in tables:
    output_rows = []
    for table_row in table.findAll('tr'):
        columns = table_row.findAll('td')
        output_row = []
        for column in columns:
            output_row.append(column.text.strip())
        if not output_row:
            output_row = ['Transaction Name', 'Min(In secs)', 'Avg(In secs)', '', '', '', '', 'Max(In secs)',
                          '90 Percentile(In secs)']
        filter_list_items = ['Init', 'End', 'Actions']
        iCounter = 0
        for items_list in filter_list_items:
            if any(items_list in s for s in output_row):
                iCounter += 1
            if iCounter == 0:
                ws.append(output_row)
            iRowCounter += 1

ws.delete_cols(10, 4)
ws.delete_cols(4, 4)

my_red = openpyxl.styles.colors.Color(rgb='00F9A2A2')
my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgcolor=my_red)


def is_comma(value):
    if value:
        return re.sub('[-,]', '', value)


for i in range(1, ws.max_row + 1):
    for j in range(2, ws.max_column + 1):
        cell_obj = ws.cell(row=i, column=j)
        if cell_obj.value:
            if not any(str(cell_obj.value).strip() in s for s in sample_table_headers):
                temp_value = is_comma(cell_obj.value).strip()
            if temp_value:
                if float('%2f' % float(str(temp_value).strip())) > 5:
                    cell_obj.fill = my_fill

            wb.save('Transaction_Report.xlsx')
